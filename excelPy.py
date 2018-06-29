import openpyxl
import crawling

def getname_from_excels(filename, sheetnum):
    retName = list()
    wb = openpyxl.load_workbook(filename)
    wsList = wb.get_sheet_names()
    ws = wb[wsList[sheetnum]]
    for i in ws.iter_rows(min_row = 2, max_col = 2):
        name = i[1].value
        retName.append(name)
    return retName

def write_to_excel(filename, sheetnum, datalist):
    wb = openpyxl.load_workbook(filename)
    wsList = wb.get_sheet_names()
    ws = wb[wsList[sheetnum]]
    for i in range(len(datalist)):
        for j in range(5):
            ws.cell(row=i + 2, column=j + 21).value = datalist[i][j]
    wb.save(filename)